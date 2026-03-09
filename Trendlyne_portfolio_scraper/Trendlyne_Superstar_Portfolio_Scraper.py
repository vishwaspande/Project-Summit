"""
Trendlyne Superstar Portfolio Scraper
=====================================
Scrapes individual and institutional investor portfolios from Trendlyne
and creates two Excel files with one tab per investor.

Requirements:
    pip install requests beautifulsoup4 openpyxl

Usage:
    python trendlyne_portfolio_scraper.py

Output:
    1. Individual_Investors_Portfolios.xlsx
    2. Institutional_Investors_Portfolios.xlsx

Data Source: Trendlyne (based on BSE/NSE exchange shareholding filings)
"""

import requests
from bs4 import BeautifulSoup
import re
import time
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONFIGURATION — Add/remove investors here
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

INDIVIDUAL_INVESTORS = {
    "Rakesh Jhunjhunwala & Associates": "https://trendlyne.com/portfolio/superstar-shareholders/53781/latest/rakesh-jhunjhunwala-and-associates-portfolio/",
    "Dolly Khanna": "https://trendlyne.com/portfolio/superstar-shareholders/53757/latest/dolly-khanna-portfolio/",
    "Ashish Kacholia": "https://trendlyne.com/portfolio/superstar-shareholders/53746/latest/ashish-kacholia-portfolio/",
    "Vijay Kedia": "https://trendlyne.com/portfolio/superstar-shareholders/53789/latest/vijay-kedia-portfolio/",
    "Radhakishan Damani": "https://trendlyne.com/portfolio/superstar-shareholders/53775/latest/radhakishan-damani-portfolio/",
    "Mukul Agrawal": "https://trendlyne.com/portfolio/superstar-shareholders/53769/latest/mukul-agrawal-portfolio/",
    "Sunil Singhania": "https://trendlyne.com/portfolio/superstar-shareholders/53785/latest/sunil-singhania-portfolio/",
    "Porinju Veliyath": "https://trendlyne.com/portfolio/superstar-shareholders/53774/latest/porinju-veliyath-portfolio/",
    "Nemish S Shah": "https://trendlyne.com/portfolio/superstar-shareholders/53771/latest/nemish-s-shah-portfolio/",
    "Mohnish Pabrai": "https://trendlyne.com/portfolio/superstar-shareholders/53768/latest/mohnish-pabrai-portfolio/",
    "Ajay Upadhyaya": "https://trendlyne.com/portfolio/superstar-shareholders/53739/latest/ajay-upadhyaya-portfolio/",
    "Anil Kumar Goel": "https://trendlyne.com/portfolio/superstar-shareholders/53741/latest/anil-kumar-goel-portfolio/",
    "Rekha Jhunjhunwala": "https://trendlyne.com/portfolio/superstar-shareholders/53782/latest/rekha-jhunjhunwala-portfolio/",
}

INSTITUTIONAL_INVESTORS = {
    "Nalanda India Fund": "https://trendlyne.com/portfolio/superstar-shareholders/53966/latest/nalanda-india-fund-limited/",
    "Government Pension Fund Global": "https://trendlyne.com/portfolio/superstar-shareholders/53957/latest/government-pension-fund-global/",
    "Elara India Opportunities Fund": "https://trendlyne.com/portfolio/superstar-shareholders/53948/latest/elara-india-opportunities-fund-limited/",
    "Amansa Holdings": "https://trendlyne.com/portfolio/superstar-shareholders/53932/latest/amansa-holdings-private-limited/",
    "East Bridge Capital Master Fund": "https://trendlyne.com/portfolio/superstar-shareholders/53947/latest/east-bridge-capital-master-fund-limited/",
    "Steadview Capital Mauritius": "https://trendlyne.com/portfolio/superstar-shareholders/53985/latest/steadview-capital-mauritius-limited/",
    "Baron Emerging Markets Fund": "https://trendlyne.com/portfolio/superstar-shareholders/53936/latest/baron-emerging-markets-fund/",
    "Abu Dhabi Investment Authority": "https://trendlyne.com/portfolio/superstar-shareholders/53929/latest/abu-dhabi-investment-authority/",
    "Societe Generale": "https://trendlyne.com/portfolio/superstar-shareholders/53983/latest/societe-generale/",
    "Vanguard Intl Stock Index Fund": "https://trendlyne.com/portfolio/superstar-shareholders/53993/latest/vanguard-total-international-stock-index-fund/",
}

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
}

DELAY_BETWEEN_REQUESTS = 2  # seconds — be polite to the server


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SCRAPING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def fetch_page(url, investor_name):
    """Fetch a Trendlyne portfolio page."""
    print(f"  Fetching: {investor_name}...", end=" ", flush=True)
    try:
        session = requests.Session()
        resp = session.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        print(f"OK ({len(resp.text):,} bytes)")
        return resp.text
    except Exception as e:
        print(f"FAILED: {e}")
        return None


def parse_portfolio(html, investor_name):
    """Parse the holdings table from Trendlyne HTML."""
    soup = BeautifulSoup(html, 'html.parser')
    holdings = []

    # Extract net worth info
    net_worth = ""
    nw_match = re.search(r'net worth of over Rs\.\s*([\d,]+\.?\d*\s*Cr)', html)
    if nw_match:
        net_worth = f"Net Worth: Rs. {nw_match.group(1).strip()}"

    # Find the main holdings table
    # It typically has headers: Stock | Holding Value Qty Held | Dec 2025 Change % | Dec 2025 Holding % | ...
    tables = soup.find_all('table')

    main_table = None
    for table in tables:
        first_row = table.find('tr')
        if first_row:
            text = first_row.get_text()
            if 'Stock' in text and ('Holding' in text or 'Value' in text):
                main_table = table
                break

    if not main_table:
        # Try finding by class or structure
        for table in tables:
            headers = table.find_all('th')
            header_text = ' '.join(h.get_text() for h in headers)
            if 'Stock' in header_text:
                main_table = table
                break

    if main_table:
        holdings = parse_main_table(main_table)
    else:
        # Fallback: extract stock info from any table or div
        holdings = fallback_extract(soup)

    return holdings, net_worth


def parse_main_table(table):
    """Parse the main portfolio holdings table."""
    holdings = []
    rows = table.find_all('tr', recursive=False)
    if not rows:
        tbody = table.find('tbody')
        if tbody:
            rows = tbody.find_all('tr', recursive=False)

    # Get header columns
    header_row = rows[0] if rows else None
    if not header_row:
        return []

    header_cells = header_row.find_all(['th', 'td'])
    headers = [h.get_text(strip=True) for h in header_cells]

    # Process data rows
    for row in rows[1:]:
        cells = row.find_all(['td', 'th'], recursive=False)
        if not cells:
            continue

        # Check if this row has a stock link (outer row, not nested)
        stock_link = row.find('a', href=re.compile(r'/equity/share-holding/\d+/'))
        if not stock_link:
            # Could be inside a nested table/div — skip
            stock_link = row.find('a', href=re.compile(r'/equity/\d+/'))
        if not stock_link:
            continue

        stock_name = stock_link.get_text(strip=True)
        if not stock_name or len(stock_name) < 2:
            continue

        # Extract BSE/NSE code from URL
        code_match = re.search(r'/equity/(?:share-holding/)?\d+/([^/]+)/', stock_link.get('href', ''))
        stock_code = code_match.group(1) if code_match else ''

        row_data = {'Stock': stock_name, 'Code': stock_code}

        # Extract values from cells
        cell_texts = [c.get_text(strip=True) for c in cells]

        # Find holding value (pattern: "1,234.5 Cr" or just a number followed by Cr)
        for ct in cell_texts:
            val_match = re.match(r'^([\d,]+\.?\d*)\s*Cr$', ct)
            if val_match:
                row_data['Holding Value (Cr)'] = val_match.group(1)
                break

        # Find qty held (large plain number)
        for ct in cell_texts:
            clean = ct.replace(',', '')
            if re.match(r'^\d{4,}$', clean):
                row_data['Qty Held'] = ct
                break

        # Extract quarterly holding percentages
        # These are typically the last N cells in the row
        quarter_labels = [
            'Latest Qtr Change %', 'Latest Qtr Holding %',
            'Prev Q1 %', 'Prev Q2 %', 'Prev Q3 %', 'Prev Q4 %',
            'Prev Q5 %', 'Prev Q6 %', 'Prev Q7 %', 'Prev Q8 %'
        ]

        # Try to use actual header labels if available
        if len(headers) > 3:
            pct_cols = [h for h in headers if '%' in h or 'Change' in h]
        else:
            pct_cols = quarter_labels

        pct_values = []
        for ct in cell_texts:
            ct = ct.strip()
            if ct == '-':
                pct_values.append('-')
            elif re.match(r'^-?\d+\.?\d*$', ct):
                # Could be a percentage value (without % sign)
                pct_values.append(ct + '%')
            elif re.match(r'^-?\d+\.?\d*%$', ct):
                pct_values.append(ct)

        # Map to named columns
        actual_pct_headers = [h for h in headers if ('%' in h) or ('Change' in h and '%' in h)]
        if not actual_pct_headers:
            actual_pct_headers = quarter_labels[:len(pct_values)]

        for i, val in enumerate(pct_values):
            if i < len(actual_pct_headers):
                row_data[actual_pct_headers[i]] = val
            else:
                row_data[f'Q{i+1} %'] = val

        # Avoid duplicate stocks
        if not any(h.get('Stock') == stock_name for h in holdings):
            holdings.append(row_data)

    return holdings


def fallback_extract(soup):
    """Fallback extraction when table parsing fails."""
    holdings = []
    seen = set()

    # Find all stock links
    links = soup.find_all('a', href=re.compile(r'/equity/share-holding/\d+/'))
    if not links:
        links = soup.find_all('a', href=re.compile(r'/equity/\d+/[A-Z]'))

    for link in links:
        name = link.get_text(strip=True)
        if not name or name in seen or len(name) < 2:
            continue

        code_match = re.search(r'/equity/(?:share-holding/)?\d+/([^/]+)/', link.get('href', ''))
        code = code_match.group(1) if code_match else ''

        # Try to get data from parent row
        parent_row = link.find_parent('tr')
        row_data = {'Stock': name, 'Code': code}

        if parent_row:
            cells = parent_row.find_all(['td', 'th'])
            cell_texts = [c.get_text(strip=True) for c in cells]

            for ct in cell_texts:
                val_match = re.match(r'^([\d,]+\.?\d*)\s*Cr$', ct)
                if val_match:
                    row_data['Holding Value (Cr)'] = val_match.group(1)
                    break

            for ct in cell_texts:
                clean = ct.replace(',', '')
                if re.match(r'^\d{5,}$', clean):
                    row_data['Qty Held'] = ct
                    break

            pct_idx = 0
            for ct in cell_texts:
                ct = ct.strip()
                if re.match(r'^-?\d+\.?\d*$', ct) or ct == '-':
                    row_data[f'Holding Q{pct_idx+1} %'] = ct if ct != '-' else '-'
                    pct_idx += 1

        seen.add(name)
        holdings.append(row_data)

    return holdings


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL FORMATTING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LIGHT_FILL = PatternFill('solid', fgColor='F2F7FB')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
DATA_FONT = Font(name='Arial', size=10)
DATA_BORDER = Border(bottom=Side(style='thin', color='E0E0E0'))


def sanitize_sheet_name(name):
    name = re.sub(r'[\\/*?\[\]:]', '', name)
    return name[:31]


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = Border(bottom=Side(style='medium', color='1F4E79'))


def style_data(ws, start_row, end_row, num_cols):
    for row in range(start_row, end_row + 1):
        fill = LIGHT_FILL if (row - start_row) % 2 == 0 else WHITE_FILL
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = DATA_FONT
            cell.border = DATA_BORDER
            cell.alignment = Alignment(vertical='center')


def auto_width(ws, num_cols, num_rows, start_row=4):
    for ci in range(1, num_cols + 1):
        max_len = 0
        for ri in range(start_row, min(start_row + num_rows + 1, start_row + 100)):
            val = ws.cell(row=ri, column=ci).value
            if val:
                max_len = max(max_len, len(str(val)))
        # Also check header
        hdr = ws.cell(row=start_row, column=ci).value
        if hdr:
            max_len = max(max_len, len(str(hdr)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 35)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL GENERATION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def create_excel(investor_dict, filename, category_label):
    """Main function: fetch, parse, and write to Excel."""
    wb = Workbook()
    wb.remove(wb.active)

    results = {}

    print(f"\n{'='*60}")
    print(f"  {category_label}")
    print(f"{'='*60}")

    for investor_name, url in investor_dict.items():
        html = fetch_page(url, investor_name)
        time.sleep(DELAY_BETWEEN_REQUESTS)

        if not html:
            results[investor_name] = 0
            sheet_name = sanitize_sheet_name(investor_name)
            ws = wb.create_sheet(title=sheet_name)
            ws.cell(row=1, column=1, value=f"{investor_name}").font = Font(bold=True, name='Arial', size=13, color='1F4E79')
            ws.cell(row=3, column=1, value="Failed to fetch data.").font = Font(color='CC0000', name='Arial')
            continue

        holdings, net_worth = parse_portfolio(html, investor_name)

        # Filter out stocks where Qty Held is '-' (not current holdings)
        holdings = [h for h in holdings if h.get('Qty Held', '-') not in ('-', '', None)]

        sheet_name = sanitize_sheet_name(investor_name)
        ws = wb.create_sheet(title=sheet_name)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        ws.cell(row=1, column=1, value=f"{investor_name} — Portfolio Holdings").font = Font(
            bold=True, name='Arial', size=13, color='1F4E79')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        subtitle = f"Source: Trendlyne (BSE/NSE exchange filings) | {net_worth}" if net_worth else \
                   f"Source: Trendlyne (BSE/NSE exchange filings) | {category_label}"
        ws.cell(row=2, column=1, value=subtitle).font = Font(name='Arial', size=9, italic=True, color='666666')

        if not holdings:
            ws.cell(row=4, column=1, value="No holdings data retrieved for this investor.").font = Font(
                name='Arial', size=10, color='CC0000')
            results[investor_name] = 0
            print(f"    → 0 holdings found for {investor_name}")
            continue

        # Collect all column keys
        all_keys = []
        for h in holdings:
            for k in h.keys():
                if k not in all_keys:
                    all_keys.append(k)

        # Write headers at row 4
        for ci, key in enumerate(all_keys, 1):
            ws.cell(row=4, column=ci, value=key)
        style_header_row(ws, 4, len(all_keys))

        # Write data starting at row 5
        for ri, holding in enumerate(holdings, 5):
            for ci, key in enumerate(all_keys, 1):
                val = holding.get(key, '-')
                ws.cell(row=ri, column=ci, value=val if val else '-')

        end_row = 4 + len(holdings)
        style_data(ws, 5, end_row, len(all_keys))
        auto_width(ws, len(all_keys), len(holdings))

        # Freeze panes and auto-filter
        ws.freeze_panes = 'A5'
        ws.auto_filter.ref = f"A4:{get_column_letter(len(all_keys))}{end_row}"

        results[investor_name] = len(holdings)
        print(f"    → {len(holdings)} holdings for {investor_name}")

    # ── Summary Sheet ──
    summary = wb.create_sheet("Summary", 0)
    summary.merge_cells('A1:D1')
    summary.cell(row=1, column=1, value=f"{category_label} — Portfolio Summary").font = Font(
        bold=True, name='Arial', size=14, color='1F4E79')
    summary.row_dimensions[1].height = 35

    summary_headers = ['#', 'Investor / Institution', 'Holdings Count', 'Navigate']
    for ci, h in enumerate(summary_headers, 1):
        summary.cell(row=3, column=ci, value=h)
    style_header_row(summary, 3, len(summary_headers))

    for idx, (inv_name, count) in enumerate(results.items(), 1):
        row = 3 + idx
        sname = sanitize_sheet_name(inv_name)
        summary.cell(row=row, column=1, value=idx)
        summary.cell(row=row, column=2, value=inv_name)
        summary.cell(row=row, column=3, value=count)
        summary.cell(row=row, column=4, value=f'=HYPERLINK("#\'{sname}\'!A1", "Go →")')
        summary.cell(row=row, column=4).font = Font(name='Arial', size=10, color='0563C1', underline='single')

    style_data(summary, 4, 3 + len(results), len(summary_headers))
    summary.column_dimensions['A'].width = 6
    summary.column_dimensions['B'].width = 42
    summary.column_dimensions['C'].width = 18
    summary.column_dimensions['D'].width = 12

    wb.save(filename)
    total = sum(results.values())
    success = sum(1 for v in results.values() if v > 0)
    print(f"\n  ✓ Saved: {filename}")
    print(f"    {success}/{len(results)} investors with data, {total} total holdings\n")
    return filename


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

if __name__ == '__main__':
    print("=" * 60)
    print("  Trendlyne Superstar Portfolio Scraper")
    print("  Data Source: BSE/NSE Exchange Shareholding Filings")
    print("=" * 60)

    f1 = create_excel(
        INDIVIDUAL_INVESTORS,
        "Individual_Investors_Portfolios.xlsx",
        "Individual Investors"
    )

    f2 = create_excel(
        INSTITUTIONAL_INVESTORS,
        "Institutional_Investors_Portfolios.xlsx",
        "Institutional Investors"
    )

    print("=" * 60)
    print("  ALL DONE!")
    print(f"  Files created in: {os.getcwd()}")
    print(f"    1. Individual_Investors_Portfolios.xlsx")
    print(f"    2. Institutional_Investors_Portfolios.xlsx")
    print("=" * 60)